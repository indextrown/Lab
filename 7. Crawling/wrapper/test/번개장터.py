import importlib
import subprocess
import os
import time
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.common.by import By
import pandas as pd
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows

# os.system('pip install --upgrade selenium')
# print(webdriver.__version__)

# 모듈 설치
def ensure_required_modules_installed():
    # 필요한 모듈 리스트
    required_modules = [
        "requests",
        "bs4",
        "selenium",
        "webdriver_manager",
        "chromedriver_autoinstaller"
        # ... add more modules if needed
    ]
    
    border_line = "#" * 64  # 64개의 #을 사용하여 직사각형 형태 생성
    print(border_line)
    print("#                필요한 패키지 확인합니다. . .                 #")
    
    max_module_length = max(len(module) for module in required_modules)
    for module in required_modules:
        try:
            importlib.import_module(module)
            spaces_after_module = " " * (max_module_length - len(module))
            print("# {} 모듈을 성공적으로 임포트했습니다.{} #".format(module, spaces_after_module))
        except ImportError:
            install_message = "'{}' 모듈이 설치되어 있지 않습니다. Installing now...".format(module)
            spaces_after_module = " " * (max_module_length - len(module))
            print("# {}{} #".format(install_message, " " * (50 - len(install_message)) + spaces_after_module))
            try:
                subprocess.check_call(["pip", "install", module])
                success_message = "'{}' 모듈 설치에 성공했습니다.".format(module)
                spaces_after_module = " " * (max_module_length - len(module))
                print("# {}{} #".format(success_message, " " * (50 - len(success_message)) + spaces_after_module))
            except Exception as e:
                error_message = "'{}' 모듈 설치 중 에러 발생: {}".format(module, e)
                spaces_after_module = " " * (max_module_length - len(module))
                print("# {}{} #".format(error_message, " " * (50 - len(error_message)) + spaces_after_module))
    
    print(border_line)

# 함수 호출하여 모듈 설치 확인
ensure_required_modules_installed()


# 크롬 드라이버 자동 업데이트
chrome_options = Options()

# user agent
User_Agent =  "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/605.1.15 (KHTML, like Gecko) Version/16.2 Safari/605.1.15",
chrome_options.add_argument(f'user-agent={User_Agent}')

chrome_options.add_argument("--headless")  # headless 모드 활성화
chrome_options.add_experimental_option("detach", True)

# 불필요한 에러 메시지 없애기
chrome_options.add_experimental_option("excludeSwitches", ["enable-logging"])

# ChromeDriverManager를 통해 크롬 드라이버를 최신 버전으로 자동 설치하여 Service 생성 및 저장
service = Service(executable_path=ChromeDriverManager().install())
driver = webdriver.Chrome(service=service, options=chrome_options)

# 웹페이지로 이동하고 웹페이지가 로딩될 때까지 최대 10초 대기
driver.implicitly_wait(10)

def mysearch(search, keyword, min_money, max_money, page):
    try:
        # 데이터를 저장할 빈 리스트
        cnt = 1
        while True:
            if cnt>page:
                break
            print(f"{cnt}페이지 입니다.")
            url = f"https://m.bunjang.co.kr/search/products?order=score&page={cnt}&q={search}"
            driver.get(url)
            li = driver.find_elements(By.CSS_SELECTOR, ".sc-dxZgTM.xRBLa")
            # time.sleep(3)
            for i in li:
                name = i.find_element(By.CSS_SELECTOR, "a>div:nth-child(2)>div").text
                price = int(i.find_element(By.CSS_SELECTOR, "a>div:nth-child(2)>div:nth-child(2)>div").text.replace(",", ""))
                url = i.find_element(By.CSS_SELECTOR, "a").get_attribute('href')
                if max_money >= price and min_money <=price:
                    if all(key.lower() in name.lower() for key in keyword):
                        print(f"제목: {name}\n가격: {price}\nURL: {url}\n")
                        data.append({"제목": name, "가격": price, "URL": url})
            cnt += 1
        return data
    
    except Exception as e:
        # 브라우저 닫기
        print("err: ", e)
        
    finally:
        driver.quit()

def data_to_xlsx(excel_filename):
    # 데이터를 pandas DataFrame으로 변환
        df = pd.DataFrame(data)
        with pd.ExcelWriter(excel_filename, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name='Sheet1')
            worksheet = writer.sheets['Sheet1']
            
            # URL을 하이퍼링크로 변환하기 위한 수정
            for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), 1):
                if r_idx == 1:  # 첫 번째 행은 헤더이므로 건너뜀
                    continue
                url = row[df.columns.get_loc("URL")]
                cell = worksheet.cell(row=r_idx, column=df.columns.get_loc("URL") + 1)
                cell.value = '=HYPERLINK("{}", "{}")'.format(url, url)

            # 열 너비 설정을 위한 코드 수정
            for column_cells in worksheet.columns:
                length = max(len(str(cell.value)) for cell in column_cells)
                col_letter = column_cells[0].column_letter
                if col_letter == worksheet.cell(row=1, column=df.columns.get_loc("URL") + 1).column_letter:
                    worksheet.column_dimensions[col_letter].width = 120  # URL 열의 너비를 20으로 설정
                elif col_letter == worksheet.cell(row=1, column=df.columns.get_loc("가격") + 1).column_letter:
                    worksheet.column_dimensions[col_letter].width = 10  # "가격" 열의 너비를 20으로 설정
                else:
                    worksheet.column_dimensions[col_letter].width = length

        
        # df.to_excel(excel_filename, index=False)
        print(f"{excel_filename}에 저장되었습니다.")
        

if __name__ == "__main__":
    search = "맥북m1"
    keyword = ["m1", "16", "512"]
    min_money = 950000
    max_money = 1300000
    page = 10
    save = True
    data = []

    
    mysearch(search, keyword, min_money, max_money, page)
    if save==True:
        data_to_xlsx('test.xlsx')

